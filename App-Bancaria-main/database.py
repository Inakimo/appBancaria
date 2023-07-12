
# importaciones
from hashlib import new
import names
import pandas as pd
from random import randint as r
from datetime import date


# salida1
import pandas as pd
from openpyxl import load_workbook

# salida errores
fecha = date.today()
sucursal = input("Ingrese el nombre de la sucursal:")


def salida_errores():

    # creacion de documento

    def typeDocument():
        errortype = r(0, 100)
        if errortype < 80:
            value = r(0, 1)
            if value == 0:
                type = '011'
            else:
                type = '099'
            return type
        if errortype > 80:
            invaliddoc = r(12, 98)
            strinvalid = "0"+str(invaliddoc)
            type = strinvalid
            return type

    # creacion de nombres aleaotorios
    def randomName():
        blankname = r(0, 100)
        if blankname < 80:
            value = r(0, 1)
            if value == 1:
                sex = 'male'
            else:
                sex = 'female'

            name = names.get_full_name(gender=sex)
        else:
            name = ""
        return name

    # seleccion de tipo de operacion realizada por cada cliente
    def typeOperation():

        value = r(1, 43)
        if value >= 1 and value <= 37:
            if len(str(value)) == 2:
                operation = '0000' + str(value)
            else:
                operation = '00000' + str(value)

        else:
            operation = "0000" + str(value)
        return operation

# creacion de documento de acuerdo el tipo de documento que tengas

    def randomDocument(valueTypeDocument):
        error = r(0, 100)
        if error < 85:
            unSexDocument = str(r(12000000, 44000000))
            if valueTypeDocument == '011':
                secondValue = r(0, 1)
                if secondValue == 0:
                    thirdValue = r(0, 1)
                    if thirdValue == 0:
                        secondType = '30'
                        document = valueTypeDocument+secondType+unSexDocument
                        return document
                    else:
                        secondType = '20'
                        document = valueTypeDocument+secondType+unSexDocument
                        return document
                elif secondValue == 1:
                    thirdValue = r(0, 1)
                    if thirdValue == 0:
                        secondType = '33'
                        document = valueTypeDocument+secondType+unSexDocument
                        return document
                    else:
                        secondType = '27'
                        document = valueTypeDocument+secondType+unSexDocument
                        return document
            else:
                document = str(valueTypeDocument)+unSexDocument
                return document
        else:
            document = 0
            return document

    # seleccion del codigo de garantia aleatorio que tendra el prestamo

    def randomGuaranteeCode():
        guaranteeCode = r(0, 37)
        return guaranteeCode

    # asigo el tipo de garantia de acuerdo al codigo de garantia ya establecido

    def typeGuarantee(guaranteeCode):
        if guaranteeCode < 33:
            if guaranteeCode == 0:
                typeValueGuarantee = 0
            elif guaranteeCode == 1 or guaranteeCode == 2 or guaranteeCode == 3 or guaranteeCode == 4 or guaranteeCode == 5 or guaranteeCode == 6 or guaranteeCode == 7 or guaranteeCode == 8 or guaranteeCode == 9 or guaranteeCode == 10 or guaranteeCode == 11 or guaranteeCode == 25 or guaranteeCode == 31:
                typeValueGuarantee = 2
            else:
                typeValueGuarantee = 1

        else:
            typeValueGuarantee = r(3, 10)
        return typeValueGuarantee

    # seleccion del codigo de moneda

    def randomCurrencyCode():
        currencyCode = r(0, 3)
        return currencyCode

    # seleccion aleatoria de tipo de Cartera

    def randomTypeWallet():
        typeWallet = r(1, 3)
        return typeWallet

    # creacion de capital de operacion de acuerdo al tipo de cartera del cliente y la moneda

    def randomOperationCapital(typeWallet, currencyCode):
        docazar = r(0, 100)
        if docazar < 95:
            if typeWallet == 1:
                operationCapital = r(50000, 30000000)
                if currencyCode == 0:
                    return operationCapital
                elif currencyCode == 1:
                    return operationCapital//300

            elif typeWallet == 2:
                operationCapital = r(100000000, 500000000)
                if currencyCode == 0:
                    return operationCapital
                elif currencyCode == 1:
                    return operationCapital//300

            elif typeWallet == 3:
                operationCapital = r(30000000, 100000000)
                if currencyCode == 0:
                    return operationCapital
                elif currencyCode == 1:
                    return operationCapital//300
        else:
            operationCapital = r(1, 100)
        return operationCapital

    # asignacion de numero de operacion

    def operationNumber(operation):
        operationModify = operation[((len(operation))-2):(len(operation))]
        if operationModify == '01':
            opNumber = '131709'
        elif operationModify == '02':
            opNumber = '13712'
        elif operationModify == '03':
            opNumber = '13712'
        elif operationModify == '04':
            opNumber = '131718'
        elif operationModify == '05' or operationModify == '21':
            opNumber = '131708'
        elif operationModify == '06':
            opNumber = '131711'
        elif operationModify == '07':
            opNumber = '131713'
        elif operationModify == '08':
            opNumber = '131714'
        elif operationModify == '09':
            opNumber = '131731'
        elif operationModify == '10' or operationModify == '11' or operationModify == '12' or operationModify == '13' or operationModify == '14' or operationModify == '15' or operationModify == '16' or operationModify == '17':
            opNumber = '131742'
        elif operationModify == '18':
            opNumber = '131741'
        elif operationModify == '19':
            opNumber = '131738'
        elif operationModify == '20':
            opNumber = '131736'
        elif operationModify == '22':
            opNumber = '132735'
        elif operationModify == '23':
            opNumber = '721731'
        elif operationModify == '24':
            opNumber = '141701'
        elif operationModify == '25':
            opNumber = '150720'
        elif operationModify == '26':
            opNumber = '171131'
        elif operationModify == '27':
            opNumber = '131101'
        elif operationModify == '28':
            opNumber = '721735'
        elif operationModify == '30':
            opNumber = '131728'
        elif operationModify == '31':
            opNumber = '135799'
        elif operationModify == '32':
            opNumber = '161003'
        elif operationModify == '33':
            opNumber = '131744'
        elif operationModify == '34':
            opNumber = '131752'
        elif operationModify == '35':
            opNumber = '131748'
        elif operationModify == '36':
            opNumber = '131792'
        else:
            opNumber = '131792'
        return opNumber

    # asignacion de situacion de deudor de acuerdo al atraso

    def debtorAssignment():
        debt = r(0, 500)
        if debt >= 0 and debt <= 30:
            debtSituation = 1
        elif debt >= 31 and debt <= 90:
            debtSituation = 2
        elif debt >= 90 and debt <= 180:
            debtSituation = 3
        elif debt >= 181 and debt <= 360:
            debtSituation = 4
        elif debt >= 361 and debt <= 460:
            debtSituation = 5
        else:
            debtSituation = r(6, 10)
        return debtSituation

    # asignacion de interes a cobrar de acuerdo al capital

    def interestCharge(typeWallet, operationCapital):
        if typeWallet == 1:
            interest = round(operationCapital*0.1)
        elif typeWallet == 2:
            interest = round(operationCapital*0.2)
        elif typeWallet == 3:
            interest = round(operationCapital*0.25)
        elif typeWallet == 4:
            interest = round(operationCapital*0.3)
        else:
            interest = round(operationCapital*0.35)
        return interest

    # creacion del dataframe
    keys = ['tipoDocumento', 'documento', 'nombreCompleto', 'operacion', 'codigoGarantia', 'tipoGarantia',
            'codigoMoneda', 'capitalOperacion', 'interesCobrar', 'clasificacionDeudor', 'numeroOperacion', 'tipoCartera']
    dataframe = []
    clients = 30
    for x in range(clients):
        # creacion de diccionario con valores independientes
        dictionary = {'tipoDocumento': typeDocument(), 'documento': 0, 'nombreCompleto': randomName(), 'operacion': typeOperation(), 'codigoGarantia': randomGuaranteeCode(), 'capitalOperacion': 0,
                      'interesCobrar': 0, 'tipoGarantia': 0, 'codigoMoneda': randomCurrencyCode(), 'clasificacionDeudor': debtorAssignment(), 'numeroOperacion': 0, 'tipoCartera': randomTypeWallet()}
        # asignando valores dependientes de los diccionarios
        dictionary['documento'] = randomDocument(dictionary['tipoDocumento'])
        dictionary['tipoGarantia'] = typeGuarantee(
            dictionary['codigoGarantia'])
        dictionary['capitalOperacion'] = randomOperationCapital(
            dictionary['tipoCartera'], dictionary['codigoMoneda'])
        dictionary['interesCobrar'] = interestCharge(
            dictionary['tipoCartera'], dictionary['capitalOperacion'])
        dictionary['numeroOperacion'] = operationNumber(
            dictionary['operacion'])
        # los coloco en el dataframe
        dataframe.append(dictionary)
        dictionary = {}

    # creacion de la base de datos
    print(dataframe)
    database = pd.DataFrame(dataframe)
    print(database)

    # creacion del archivo excel
    database.to_excel('database.xlsx')

    # Suma intereses a cobrar
    h = 0
    sumaintereses = 0
    for h in range(len(database)):
        sumaintereses = sumaintereses+database.iloc[h]['interesCobrar']
        h+1

    k = 0
    sumacapital = 0
    for k in range(len(database)):
        sumacapital = sumacapital+database.iloc[k]['capitalOperacion']
        k+1

    errores = []

    # errortipodocumento
    j = 0
    errortipodocumento = 0
    capitalerrortipodoc = 0
    for j in range(len(database)):
        if database.iloc[j]['tipoDocumento'] != '011' and database.iloc[j]['tipoDocumento'] != '099':
            errortipodocumento = errortipodocumento+1
            if j not in errores:
                capitalerrortipodoc = capitalerrortipodoc + \
                    database.iloc[j]['capitalOperacion']
                errores.append(j)
        j+1

    # errordocumentoenblanco
    i = 0
    capitalerrordocumento = 0
    errordocumento = 0
    for i in range(len(database)):
        if database.iloc[i]['documento'] == 0:
            errordocumento = errordocumento+1
            if i not in errores:
                capitalerrordocumento = capitalerrordocumento + \
                    database.iloc[i]['capitalOperacion']
                errores.append(i)
        i+1

    # error nombre en blanco
    a = 0
    errornombre = 0
    capitalerrornombre = 0
    for a in range(len(database)):
        if database.iloc[a]['nombreCompleto'] == "":
            errornombre = errornombre+1
            if a not in errores:
                capitalerrornombre = capitalerrornombre + \
                    database.iloc[a]['capitalOperacion']
                errores.append(a)
        a+1

    # error operacion
    b = 0
    erroroperacion = 0
    capitalerroroperacion = 0
    for b in range(len(database)):
        if database.iloc[b]['operacion'] != '000001' and database.iloc[b]['operacion'] != '000002' and database.iloc[b]['operacion'] != '000003' and database.iloc[b]['operacion'] != '000004' and database.iloc[b]['operacion'] != '000005' and database.iloc[b]['operacion'] != '000006' and database.iloc[b]['operacion'] != '000007' and database.iloc[b]['operacion'] != '000008' and database.iloc[b]['operacion'] != '000009' and database.iloc[b]['operacion'] != '000010' and database.iloc[b]['operacion'] != '000011' and database.iloc[b]['operacion'] != '000012' and database.iloc[b]['operacion'] != '000013' and database.iloc[b]['operacion'] != '000014' and database.iloc[b]['operacion'] != '000015' and database.iloc[b]['operacion'] != '000016' and database.iloc[b]['operacion'] != '000017' and database.iloc[b]['operacion'] != '000018' and database.iloc[b]['operacion'] != '000019' and database.iloc[b]['operacion'] != '000020' and database.iloc[b]['operacion'] != '000021' and database.iloc[b]['operacion'] != '000022' and database.iloc[b]['operacion'] != '000023' and database.iloc[b]['operacion'] != '000024' and database.iloc[b]['operacion'] != '000025' and database.iloc[b]['operacion'] != '000026' and database.iloc[b]['operacion'] != '000027' and database.iloc[b]['operacion'] != '000028' and database.iloc[b]['operacion'] != '000029' and database.iloc[b]['operacion'] != '000030' and database.iloc[b]['operacion'] != '000031' and database.iloc[b]['operacion'] != '000032' and database.iloc[b]['operacion'] != '000033' and database.iloc[b]['operacion'] != '000034' and database.iloc[b]['operacion'] != '000035' and database.iloc[b]['operacion'] != '000036' and database.iloc[b]['operacion'] != '000037':
            erroroperacion = erroroperacion+1
            if b not in errores:
                capitalerroroperacion = capitalerroroperacion + \
                    database.iloc[b]['capitalOperacion']
                errores.append(b)
        b+1

    # codigo de garantia error
    c = 0
    errorcodgarantia = 0
    caperrorcodgarantia = 0
    for c in range(len(database)):
        if database.iloc[c]['codigoGarantia'] > 33:
            errorcodgarantia = errorcodgarantia+1
            if c not in errores:
                caperrorcodgarantia = caperrorcodgarantia + \
                    database.iloc[c]['capitalOperacion']
                errores.append(c)
        c+1

    # tipo de garantia error
    d = 0
    errortipogarantia = 0
    captipogarantia = 0
    for d in range(len(database)):
        if database.iloc[d]['tipoGarantia'] >= 3:
            errortipogarantia = errortipogarantia+1
            if d not in errores:
                captipogarantia = captipogarantia + \
                    database.iloc[d]['capitalOperacion']
                errores.append(d)
        d+1

    # codigo de moneda no definido
    e = 0
    errorcodmoneda = 0
    caperrorcodmoneda = 0
    for e in range(len(database)):
        if database.iloc[e]['codigoMoneda'] > 2:
            errorcodmoneda = errorcodmoneda+1
            if e not in errores:
                caperrorcodmoneda = caperrorcodmoneda + \
                    database.iloc[e]['capitalOperacion']
                errores.append(e)
        e+1

    # capitaldeoperacion no valido
    f = 0
    errorcapoperacion = 0
    for f in range(len(database)):
        if database.iloc[f]['capitalOperacion'] == 0:
            errorcapoperacion = errorcapoperacion+1
        f+1

    # clasificacionDeudor no definida
    g = 0
    errorclasificaciondedeudor = 0
    caperrorcladeudor = 0
    for g in range(len(database)):
        if database.iloc[g]['clasificacionDeudor'] > 5:
            errorclasificaciondedeudor = errorclasificaciondedeudor+1
            if g not in errores:
                caperrorcladeudor = caperrorcladeudor + \
                    database.iloc[g]['capitalOperacion']
                errores.append(g)
        g+1

    data = ["Tipo de documento invalido ", "Numero de documento invalido ", "Nombre en blanco ", "Codigo de operacion",
            "Codigo de garantia", "Tipo de garantia", "Codigo de moneda", "Capital no valido", "Clasificacion de deudor"]
    columns = ["Tipo de error", "Cantidad de errores", " ", "  ",
               "   ", "    ", "     ", "      ", "Suma de capital"]

    dbs1 = []
    dfs1 = pd.DataFrame(dbs1)
    dfs1.to_excel('controlerrores.xlsx')

    filesheet = "./controlerrores.xlsx"
    wb = load_workbook(filesheet)

    sheet = wb.active

    sheet.merge_cells('A1:L1')
    sheet['A1'] = "Control de errores"
    sheet['A3'] = "Sucursal"
    sheet['A5'] = "Fecha de la cartera"
    sheet['A7'] = "Total del capital de las operaciones (INT)"
    sheet['A8'] = "Interes devengado a cobrar(INT)"
    sheet['A9'] = "Deuda total Cap+Intc(INT)"
    sheet['A10'] = "Cantidad de registros del archivo"
    sheet['H3'] = sucursal
    sheet['H5'] = fecha
    sheet['H7'] = str(sumacapital)+"$"
    sheet['H8'] = str(sumaintereses)+"$"
    sheet['H9'] = str(sumacapital+sumaintereses)+"$"
    sheet['H10'] = clients

    sheet.merge_cells('A11:L11')
    sheet['A11'] = "--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
    sheet['A12'] = "Tipo de error"
    sheet['A14'] = "Tipo de documento invalido-"
    sheet['A16'] = "Numero de documento invalido-"
    sheet['A18'] = "Nombre en blanco-"
    sheet['A20'] = "Codigo de operacion invalido-"
    sheet['A22'] = "Codigo de garantia-"
    sheet['A24'] = "Tipo de garantia-"
    sheet['A26'] = "Codigo de moneda-"
    sheet['A28'] = "Capital no valido-"
    sheet['A30'] = "Clasificacion de deudor-"
    sheet['B12'] = "Cantidad de errores"
    sheet['B14'] = errortipodocumento
    sheet['B16'] = errordocumento
    sheet['B18'] = errornombre
    sheet['B20'] = erroroperacion
    sheet['B22'] = errorcodgarantia
    sheet['B24'] = errortipogarantia
    sheet['B26'] = errorcodmoneda
    sheet['B28'] = errorcapoperacion
    sheet['B30'] = errorclasificaciondedeudor
    sheet['H12'] = "Suma capital de errores"
    sheet['H14'] = str(capitalerrortipodoc)+"$"
    sheet['H16'] = str(capitalerrordocumento)+"$"
    sheet['H18'] = str(capitalerrornombre)+"$"
    sheet['H20'] = str(capitalerroroperacion)+"$"
    sheet['H22'] = str(caperrorcodgarantia)+"$"
    sheet['H24'] = str(captipogarantia)+"$"
    sheet['H26'] = str(caperrorcodmoneda)+"$"
    sheet['H30'] = str(caperrorcladeudor)+"$"

    wb.save(filesheet)

# salida 1


def salida_uno(clients):

    # empiezo leyendo el excel general con este comando
    excel = load_workbook('database.xlsx', data_only=True)
    # eligo la hoja del archivo de excel que quiero usar
    sheet_excel = excel.active
    dic = {}
    # capturo del excel toda la columna de capital de operacion
    capital_operacion = sheet_excel['G2': 'G31']
    # capturo del excel toda la columna de clasificacion de dedor
    clasificacionDeudor = sheet_excel['K2': 'K31']
    # capturo del excel toda la columna de tipo de garantias
    tipoGarantia = sheet_excel['I2': 'I31']
    # capturo del excel toda la columna de tipo de cartera
    tipoCartera = sheet_excel['M2': 'M31']

    # recorro la columna de capital de operacion hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_capitalOperacion = []
    for fila in capital_operacion:
        for celda in fila:
            lst_capitalOperacion.append(celda.value)
    # recorro la columna de capital de clasificacion deudor hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_clasificacionDeudor = []
    for fila in clasificacionDeudor:
        for celda in fila:
            lst_clasificacionDeudor.append(celda.value)
    # recorro la columna de tipo de garantia hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_tipoGarantia = []
    for fila in tipoGarantia:
        for celda in fila:
            lst_tipoGarantia.append(celda.value)
    # recorro la columna de tipo de cartera hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_tipoCartera = []
    for fila in tipoCartera:
        for celda in fila:
            lst_tipoCartera.append(celda.value)

    # creo una lista para guardar los diccionarios los cuales van a contener la informacion obtenida de las listas para cada cliente
    lst_dicAux = []  # creacion de lista vacia
    for x in range(clients):
        # creacion de los diccionarios vacios
        lst_dicAux.append({'tipoCartera': 0, 'situacionDeudor': 0,
                           'tipoGarantia': 0, 'capital': 0})

    aux = 0
    aux2 = 0
    aux3 = 0
    aux4 = 0
    # recorro los dicconarios vacios
    for dic in lst_dicAux:
        for keys in dic:  # recorro las keys de dicconario por dicconario
            if keys == 'tipoCartera':  # busco asignar los valores a la key tipoCartera
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_tipoCartera[aux]
                aux += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'situacionDeudor':  # busco asignar los valores a la key situacionDeudor
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_clasificacionDeudor[aux2]
                aux2 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'tipoGarantia':  # busco asignar los valores a la key tipoGarantia
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_tipoGarantia[aux3]
                aux3 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'capital':  # busco asignar los valores a la key capital
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_capitalOperacion[aux4]
                aux4 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel

    # una vez agrupado todos los datos obtenidos en la lista de dicconario, empiezo a filtrar, es decir, utilizar solo los que tienen tipoCartera=1 ya que esta salida requiere solo esos
    # lista donde voy a guardar los diccionarios que tienen tipo de cartera 1
    lst_carteraConsumo = []
    for dic in lst_dicAux:  # recorro los diccionarios de la lista lst_dicAux
        for keys in dic:
            if keys == 'tipoCartera':  # busco analizar la key tipoCartera
                if dic['tipoCartera'] == 1:  # busco solo los que tegna tipoCartera=1
                    # guardo los dicconarios de tipoCartera=1
                    lst_carteraConsumo.append(dic)

    # Creo las siguientes lista para guardar los diccionarios de acuerdo a la situacionDeudor
    lst_deudor1 = []
    lst_deudor2 = []
    lst_deudor3 = []
    lst_deudor4 = []
    lst_deudor5 = []

    for dic in lst_carteraConsumo:  # recorro los dicconarios de la lista de cartera de consumos creada antes
        for keys in dic:  # recorro las keys de diccionario por diccionario
            if keys == 'situacionDeudor':  # analizo la key situacionDeudor
                if dic['situacionDeudor'] == 1:  # busco los que tienen situacionDeudor=1
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor1.append(dic)
                if dic['situacionDeudor'] == 2:  # busco los que tienen situacionDeudor=2
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor2.append(dic)
                if dic['situacionDeudor'] == 3:  # busco los que tienen situacionDeudor=3
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor3.append(dic)
                if dic['situacionDeudor'] == 4:  # busco los que tienen situacionDeudor=4
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor4.append(dic)

    '''------------------------ DEUDA 1 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 1, de acuerdo al tipo de garantias que tenga
    dic_sd1g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd1g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd1g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda1 = 0
    suma_capital1_deuda1 = 0
    suma_capital2_deuda1 = 0

    # recorra la lista de los de situacion de deuda 1
    for dic in lst_deudor1:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda1 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda1 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda1 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd1g0['capitalTotal'] = suma_capital0_deuda1
    dic_sd1g1['capitalTotal'] = suma_capital1_deuda1
    dic_sd1g2['capitalTotal'] = suma_capital2_deuda1

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor1 = [0, 0, 0]
    lst_deudor1[0] = dic_sd1g0
    lst_deudor1[1] = dic_sd1g1
    lst_deudor1[2] = dic_sd1g2

    # print(lst_deudor1)

    '''------------------------ DEUDA 2 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 2, de acuerdo al tipo de garantias que tenga
    dic_sd2g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd2g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd2g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda2 = 0
    suma_capital1_deuda2 = 0
    suma_capital2_deuda2 = 0

    # recorra la lista de los de situacion de deuda 2
    for dic in lst_deudor2:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda2 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda2 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda2 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd2g0['capitalTotal'] = suma_capital0_deuda2
    dic_sd2g1['capitalTotal'] = suma_capital1_deuda2
    dic_sd2g2['capitalTotal'] = suma_capital2_deuda2

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor2 = [0, 0, 0]
    lst_deudor2[0] = dic_sd2g0
    lst_deudor2[1] = dic_sd2g1
    lst_deudor2[2] = dic_sd2g2

    # print(lst_deudor2)

    '''------------------------ DEUDA 3 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 3, de acuerdo al tipo de garantias que tenga
    dic_sd3g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd3g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd3g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda3 = 0
    suma_capital1_deuda3 = 0
    suma_capital2_deuda3 = 0

    # recorra la lista de los de situacion de deuda 3
    for dic in lst_deudor3:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda3 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda3 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda3 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd3g0['capitalTotal'] = suma_capital0_deuda3
    dic_sd3g1['capitalTotal'] = suma_capital1_deuda3
    dic_sd3g2['capitalTotal'] = suma_capital2_deuda3

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor3 = [0, 0, 0]
    lst_deudor3[0] = dic_sd3g0
    lst_deudor3[1] = dic_sd3g1
    lst_deudor3[2] = dic_sd3g2

    # print(lst_deudor3)
    '''------------------------ DEUDA 4 ------------------------ '''

    # creo diccionarios para organizar la situacion de deuda en este caso 4, de acuerdo al tipo de garantia que tenga
    dic_sd4g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd4g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd4g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda4 = 0
    suma_capital1_deuda4 = 0
    suma_capital2_deuda4 = 0

    # recorra la lista de los de situacion de deuda 4
    for dic in lst_deudor4:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda4 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda4 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda4 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd4g0['capitalTotal'] = suma_capital0_deuda4
    dic_sd4g1['capitalTotal'] = suma_capital1_deuda4
    dic_sd4g2['capitalTotal'] = suma_capital2_deuda4

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor4 = [0, 0, 0]
    lst_deudor4[0] = dic_sd4g0
    lst_deudor4[1] = dic_sd4g1
    lst_deudor4[2] = dic_sd4g2
    # print(lst_deudor4)

    '''------------------------ DEUDA 5 ------------------------ '''

    # creo diccionarios para organizar la situacion de deuda en este caso 5, de acuerdo al tipo de garantia que tenga
    dic_sd5g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd5g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd5g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda5 = 0
    suma_capital1_deuda5 = 0
    suma_capital2_deuda5 = 0

    # recorra la lista de los de situacion de deuda 5
    for dic in lst_deudor5:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda5 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda5 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda5 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd5g0['capitalTotal'] = suma_capital0_deuda5
    dic_sd5g1['capitalTotal'] = suma_capital1_deuda5
    dic_sd5g2['capitalTotal'] = suma_capital2_deuda5

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor5 = [0, 0, 0]
    lst_deudor5[0] = dic_sd5g0
    lst_deudor5[1] = dic_sd5g1
    lst_deudor5[2] = dic_sd5g2
    # print(lst_deudor5)

    # lista data frame
    lista_deudores = []

    for x in lst_deudor1:
        lista_deudores.append(x)

    for x in lst_deudor2:
        lista_deudores.append(x)

    for x in lst_deudor3:
        lista_deudores.append(x)

    for x in lst_deudor4:
        lista_deudores.append(x)

    for x in lst_deudor5:
        lista_deudores.append(x)

    print(lista_deudores, 'lista deudores')

    # Lista para generar el excel
    datosDeudores = pd.DataFrame(lista_deudores)

    # creacion del excel
    datosDeudores.to_excel('cartera_consumo.xlsx')

    # ruta de nuestro archivo
    filesheet = "./cartera_consumo.xlsx"

    # creamos el objeto load_workbook
    wb = load_workbook(filesheet)

    # Seleccionamos el archivo
    sheet = wb.active

    # Deudor 1 - COlumna excel

    sheet['A1'] = 'Situacion deudor'
    sheet['A2'] = '1'
    sheet['A3'] = '1'
    sheet['A4'] = '1'

    sheet['A5'] = '2'
    sheet['A6'] = '2'
    sheet['A7'] = '2'

    sheet['A8'] = '3'
    sheet['A9'] = '3'
    sheet['A10'] = '3'

    sheet['A11'] = '4'
    sheet['A12'] = '4'
    sheet['A13'] = '4'

    sheet['A14'] = '5'
    sheet['A15'] = '5'
    sheet['A16'] = '5'
    # Guardamos el archivo con los cambios
    wb.save(filesheet)

def salida_dos(clients):

    # empiezo leyendo el excel general con este comando
    excel = load_workbook('database.xlsx', data_only=True)
    # eligo la hoja del archivo de excel que quiero usar
    sheet_excel = excel.active
    dic = {}
    # capturo del excel toda la columna de capital de operacion
    capital_operacion = sheet_excel['G2': 'G31']
    # capturo del excel toda la columna de clasificacion de dedor
    clasificacionDeudor = sheet_excel['K2': 'K31']
    # capturo del excel toda la columna de tipo de garantias
    tipoGarantia = sheet_excel['I2': 'I31']
    # capturo del excel toda la columna de tipo de cartera
    tipoCartera = sheet_excel['M2': 'M31']

    # recorro la columna de capital de operacion hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_capitalOperacion = []
    for fila in capital_operacion:
        for celda in fila:
            lst_capitalOperacion.append(celda.value)
    # recorro la columna de capital de clasificacion deudor hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_clasificacionDeudor = []
    for fila in clasificacionDeudor:
        for celda in fila:
            lst_clasificacionDeudor.append(celda.value)
    # recorro la columna de tipo de garantia hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_tipoGarantia = []
    for fila in tipoGarantia:
        for celda in fila:
            lst_tipoGarantia.append(celda.value)
    # recorro la columna de tipo de cartera hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_tipoCartera = []
    for fila in tipoCartera:
        for celda in fila:
            lst_tipoCartera.append(celda.value)

    # creo una lista para guardar los diccionarios los cuales van a contener la informacion obtenida de las listas para cada cliente
    lst_dicAux = []  # creacion de lista vacia
    for x in range(clients):
        # creacion de los diccionarios vacios
        lst_dicAux.append({'tipoCartera': 0, 'situacionDeudor': 0,
                           'tipoGarantia': 0, 'capital': 0})

    aux = 0
    aux2 = 0
    aux3 = 0
    aux4 = 0
    # recorro los dicconarios vacios
    for dic in lst_dicAux:
        for keys in dic:  # recorro las keys de dicconario por dicconario
            if keys == 'tipoCartera':  # busco asignar los valores a la key tipoCartera
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_tipoCartera[aux]
                aux += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'situacionDeudor':  # busco asignar los valores a la key situacionDeudor
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_clasificacionDeudor[aux2]
                aux2 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'tipoGarantia':  # busco asignar los valores a la key tipoGarantia
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_tipoGarantia[aux3]
                aux3 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'capital':  # busco asignar los valores a la key capital
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_capitalOperacion[aux4]
                aux4 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel

    # una vez agrupado todos los datos obtenidos en la lista de dicconario, empiezo a filtrar, es decir, utilizar solo los que tienen tipoCartera=1 ya que esta salida requiere solo esos
    # lista donde voy a guardar los diccionarios que tienen tipo de cartera 1
    lst_carteraConsumo = []
    for dic in lst_dicAux:  # recorro los diccionarios de la lista lst_dicAux
        for keys in dic:
            if keys == 'tipoCartera':  # busco analizar la key tipoCartera
                if dic['tipoCartera'] == 2:  # busco solo los que tegna tipoCartera=1
                    # guardo los dicconarios de tipoCartera=1
                    lst_carteraConsumo.append(dic)

    # Creo las siguientes lista para guardar los diccionarios de acuerdo a la situacionDeudor
    lst_deudor1 = []
    lst_deudor2 = []
    lst_deudor3 = []
    lst_deudor4 = []
    lst_deudor5 = []

    for dic in lst_carteraConsumo:  # recorro los dicconarios de la lista de cartera de consumos creada antes
        for keys in dic:  # recorro las keys de diccionario por diccionario
            if keys == 'situacionDeudor':  # analizo la key situacionDeudor
                if dic['situacionDeudor'] == 1:  # busco los que tienen situacionDeudor=1
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor1.append(dic)
                if dic['situacionDeudor'] == 2:  # busco los que tienen situacionDeudor=2
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor2.append(dic)
                if dic['situacionDeudor'] == 3:  # busco los que tienen situacionDeudor=3
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor3.append(dic)
                if dic['situacionDeudor'] == 4:  # busco los que tienen situacionDeudor=4
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor4.append(dic)

    '''------------------------ DEUDA 1 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 1, de acuerdo al tipo de garantias que tenga
    dic_sd1g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd1g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd1g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda1 = 0
    suma_capital1_deuda1 = 0
    suma_capital2_deuda1 = 0

    # recorra la lista de los de situacion de deuda 1
    for dic in lst_deudor1:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda1 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda1 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda1 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd1g0['capitalTotal'] = suma_capital0_deuda1
    dic_sd1g1['capitalTotal'] = suma_capital1_deuda1
    dic_sd1g2['capitalTotal'] = suma_capital2_deuda1

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor1 = [0, 0, 0]
    lst_deudor1[0] = dic_sd1g0
    lst_deudor1[1] = dic_sd1g1
    lst_deudor1[2] = dic_sd1g2

    # print(lst_deudor1)

    '''------------------------ DEUDA 2 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 2, de acuerdo al tipo de garantias que tenga
    dic_sd2g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd2g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd2g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda2 = 0
    suma_capital1_deuda2 = 0
    suma_capital2_deuda2 = 0

    # recorra la lista de los de situacion de deuda 2
    for dic in lst_deudor2:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda2 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda2 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda2 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd2g0['capitalTotal'] = suma_capital0_deuda2
    dic_sd2g1['capitalTotal'] = suma_capital1_deuda2
    dic_sd2g2['capitalTotal'] = suma_capital2_deuda2

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor2 = [0, 0, 0]
    lst_deudor2[0] = dic_sd2g0
    lst_deudor2[1] = dic_sd2g1
    lst_deudor2[2] = dic_sd2g2

    # print(lst_deudor2)

    '''------------------------ DEUDA 3 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 3, de acuerdo al tipo de garantias que tenga
    dic_sd3g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd3g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd3g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda3 = 0
    suma_capital1_deuda3 = 0
    suma_capital2_deuda3 = 0

    # recorra la lista de los de situacion de deuda 3
    for dic in lst_deudor3:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda3 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda3 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda3 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd3g0['capitalTotal'] = suma_capital0_deuda3
    dic_sd3g1['capitalTotal'] = suma_capital1_deuda3
    dic_sd3g2['capitalTotal'] = suma_capital2_deuda3

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor3 = [0, 0, 0]
    lst_deudor3[0] = dic_sd3g0
    lst_deudor3[1] = dic_sd3g1
    lst_deudor3[2] = dic_sd3g2

    # print(lst_deudor3)
    '''------------------------ DEUDA 4 ------------------------ '''

    # creo diccionarios para organizar la situacion de deuda en este caso 4, de acuerdo al tipo de garantia que tenga
    dic_sd4g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd4g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd4g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda4 = 0
    suma_capital1_deuda4 = 0
    suma_capital2_deuda4 = 0

    # recorra la lista de los de situacion de deuda 4
    for dic in lst_deudor4:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda4 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda4 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda4 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd4g0['capitalTotal'] = suma_capital0_deuda4
    dic_sd4g1['capitalTotal'] = suma_capital1_deuda4
    dic_sd4g2['capitalTotal'] = suma_capital2_deuda4

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor4 = [0, 0, 0]
    lst_deudor4[0] = dic_sd4g0
    lst_deudor4[1] = dic_sd4g1
    lst_deudor4[2] = dic_sd4g2
    # print(lst_deudor4)

    '''------------------------ DEUDA 5 ------------------------ '''

    # creo diccionarios para organizar la situacion de deuda en este caso 5, de acuerdo al tipo de garantia que tenga
    dic_sd5g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd5g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd5g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda5 = 0
    suma_capital1_deuda5 = 0
    suma_capital2_deuda5 = 0

    # recorra la lista de los de situacion de deuda 5
    for dic in lst_deudor5:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda5 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda5 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda5 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd5g0['capitalTotal'] = suma_capital0_deuda5
    dic_sd5g1['capitalTotal'] = suma_capital1_deuda5
    dic_sd5g2['capitalTotal'] = suma_capital2_deuda5

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor5 = [0, 0, 0]
    lst_deudor5[0] = dic_sd5g0
    lst_deudor5[1] = dic_sd5g1
    lst_deudor5[2] = dic_sd5g2
    # print(lst_deudor5)

    # lista data frame
    lista_deudores = []

    for x in lst_deudor1:
        lista_deudores.append(x)

    for x in lst_deudor2:
        lista_deudores.append(x)

    for x in lst_deudor3:
        lista_deudores.append(x)

    for x in lst_deudor4:
        lista_deudores.append(x)

    for x in lst_deudor5:
        lista_deudores.append(x)

    print(lista_deudores, 'lista deudores cartera comercial')

    # Lista para generar el excel
    datosDeudores = pd.DataFrame(lista_deudores)

    # creacion del excel
    datosDeudores.to_excel('cartera_comercial.xlsx')

    # ruta de nuestro archivo
    filesheet = "./cartera_comercial.xlsx"

    # creamos el objeto load_workbook
    wb = load_workbook(filesheet)

    # Seleccionamos el archivo
    sheet = wb.active

    # Deudor 1 - COlumna excel

    sheet['A1'] = 'Situacion deudor'
    sheet['A2'] = '1'
    sheet['A3'] = '1'
    sheet['A4'] = '1'

    sheet['A5'] = '2'
    sheet['A6'] = '2'
    sheet['A7'] = '2'

    sheet['A8'] = '3'
    sheet['A9'] = '3'
    sheet['A10'] = '3'

    sheet['A11'] = '4'
    sheet['A12'] = '4'
    sheet['A13'] = '4'

    sheet['A14'] = '5'
    sheet['A15'] = '5'
    sheet['A16'] = '5'
    # Guardamos el archivo con los cambios
    wb.save(filesheet)

def salida_tres(clients):

    # empiezo leyendo el excel general con este comando
    excel = load_workbook('database.xlsx', data_only=True)
    # eligo la hoja del archivo de excel que quiero usar
    sheet_excel = excel.active
    dic = {}
    # capturo del excel toda la columna de capital de operacion
    capital_operacion = sheet_excel['G2': 'G31']
    # capturo del excel toda la columna de clasificacion de dedor
    clasificacionDeudor = sheet_excel['K2': 'K31']
    # capturo del excel toda la columna de tipo de garantias
    tipoGarantia = sheet_excel['I2': 'I31']
    # capturo del excel toda la columna de tipo de cartera
    tipoCartera = sheet_excel['M2': 'M31']

    # recorro la columna de capital de operacion hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_capitalOperacion = []
    for fila in capital_operacion:
        for celda in fila:
            lst_capitalOperacion.append(celda.value)
    # recorro la columna de capital de clasificacion deudor hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_clasificacionDeudor = []
    for fila in clasificacionDeudor:
        for celda in fila:
            lst_clasificacionDeudor.append(celda.value)
    # recorro la columna de tipo de garantia hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_tipoGarantia = []
    for fila in tipoGarantia:
        for celda in fila:
            lst_tipoGarantia.append(celda.value)
    # recorro la columna de tipo de cartera hasta llegar a la celda, para obtener su valor utilizo el .value en la celda y luego lo guardo en una lista todos las celdas con sus respectivo valores
    lst_tipoCartera = []
    for fila in tipoCartera:
        for celda in fila:
            lst_tipoCartera.append(celda.value)

    # creo una lista para guardar los diccionarios los cuales van a contener la informacion obtenida de las listas para cada cliente
    lst_dicAux = []  # creacion de lista vacia
    for x in range(clients):
        # creacion de los diccionarios vacios
        lst_dicAux.append({'tipoCartera': 0, 'situacionDeudor': 0,
                           'tipoGarantia': 0, 'capital': 0})

    aux = 0
    aux2 = 0
    aux3 = 0
    aux4 = 0
    # recorro los dicconarios vacios
    for dic in lst_dicAux:
        for keys in dic:  # recorro las keys de dicconario por dicconario
            if keys == 'tipoCartera':  # busco asignar los valores a la key tipoCartera
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_tipoCartera[aux]
                aux += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'situacionDeudor':  # busco asignar los valores a la key situacionDeudor
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_clasificacionDeudor[aux2]
                aux2 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'tipoGarantia':  # busco asignar los valores a la key tipoGarantia
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_tipoGarantia[aux3]
                aux3 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel
            if keys == 'capital':  # busco asignar los valores a la key capital
                # asigno valor apartir de las listas de las columnas del excel
                dic[keys] = lst_capitalOperacion[aux4]
                aux4 += 1  # utilizo estos auxiliares para recorrer las listas de las columnas del excel

    # una vez agrupado todos los datos obtenidos en la lista de dicconario, empiezo a filtrar, es decir, utilizar solo los que tienen tipoCartera=1 ya que esta salida requiere solo esos
    # lista donde voy a guardar los diccionarios que tienen tipo de cartera 1
    lst_carteraConsumo = []
    for dic in lst_dicAux:  # recorro los diccionarios de la lista lst_dicAux
        for keys in dic:
            if keys == 'tipoCartera':  # busco analizar la key tipoCartera
                if dic['tipoCartera'] == 3:  # busco solo los que tegna tipoCartera=1
                    # guardo los dicconarios de tipoCartera=1
                    lst_carteraConsumo.append(dic)

    # Creo las siguientes lista para guardar los diccionarios de acuerdo a la situacionDeudor
    lst_deudor1 = []
    lst_deudor2 = []
    lst_deudor3 = []
    lst_deudor4 = []
    lst_deudor5 = []

    for dic in lst_carteraConsumo:  # recorro los dicconarios de la lista de cartera de consumos creada antes
        for keys in dic:  # recorro las keys de diccionario por diccionario
            if keys == 'situacionDeudor':  # analizo la key situacionDeudor
                if dic['situacionDeudor'] == 1:  # busco los que tienen situacionDeudor=1
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor1.append(dic)
                if dic['situacionDeudor'] == 2:  # busco los que tienen situacionDeudor=2
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor2.append(dic)
                if dic['situacionDeudor'] == 3:  # busco los que tienen situacionDeudor=3
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor3.append(dic)
                if dic['situacionDeudor'] == 4:  # busco los que tienen situacionDeudor=4
                    # apendeo los diccionarios en la lista correspondiente
                    lst_deudor4.append(dic)

    '''------------------------ DEUDA 1 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 1, de acuerdo al tipo de garantias que tenga
    dic_sd1g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd1g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd1g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda1 = 0
    suma_capital1_deuda1 = 0
    suma_capital2_deuda1 = 0

    # recorra la lista de los de situacion de deuda 1
    for dic in lst_deudor1:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda1 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda1 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda1 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd1g0['capitalTotal'] = suma_capital0_deuda1
    dic_sd1g1['capitalTotal'] = suma_capital1_deuda1
    dic_sd1g2['capitalTotal'] = suma_capital2_deuda1

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor1 = [0, 0, 0]
    lst_deudor1[0] = dic_sd1g0
    lst_deudor1[1] = dic_sd1g1
    lst_deudor1[2] = dic_sd1g2

    # print(lst_deudor1)

    '''------------------------ DEUDA 2 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 2, de acuerdo al tipo de garantias que tenga
    dic_sd2g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd2g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd2g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda2 = 0
    suma_capital1_deuda2 = 0
    suma_capital2_deuda2 = 0

    # recorra la lista de los de situacion de deuda 2
    for dic in lst_deudor2:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda2 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda2 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda2 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd2g0['capitalTotal'] = suma_capital0_deuda2
    dic_sd2g1['capitalTotal'] = suma_capital1_deuda2
    dic_sd2g2['capitalTotal'] = suma_capital2_deuda2

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor2 = [0, 0, 0]
    lst_deudor2[0] = dic_sd2g0
    lst_deudor2[1] = dic_sd2g1
    lst_deudor2[2] = dic_sd2g2

    # print(lst_deudor2)

    '''------------------------ DEUDA 3 ------------------------ '''

    # Diccionarios para organizaar la situacion de deduda en este caso 3, de acuerdo al tipo de garantias que tenga
    dic_sd3g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd3g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd3g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda3 = 0
    suma_capital1_deuda3 = 0
    suma_capital2_deuda3 = 0

    # recorra la lista de los de situacion de deuda 3
    for dic in lst_deudor3:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda3 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda3 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda3 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd3g0['capitalTotal'] = suma_capital0_deuda3
    dic_sd3g1['capitalTotal'] = suma_capital1_deuda3
    dic_sd3g2['capitalTotal'] = suma_capital2_deuda3

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor3 = [0, 0, 0]
    lst_deudor3[0] = dic_sd3g0
    lst_deudor3[1] = dic_sd3g1
    lst_deudor3[2] = dic_sd3g2

    # print(lst_deudor3)
    '''------------------------ DEUDA 4 ------------------------ '''

    # creo diccionarios para organizar la situacion de deuda en este caso 4, de acuerdo al tipo de garantia que tenga
    dic_sd4g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd4g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd4g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda4 = 0
    suma_capital1_deuda4 = 0
    suma_capital2_deuda4 = 0

    # recorra la lista de los de situacion de deuda 4
    for dic in lst_deudor4:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda4 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda4 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda4 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd4g0['capitalTotal'] = suma_capital0_deuda4
    dic_sd4g1['capitalTotal'] = suma_capital1_deuda4
    dic_sd4g2['capitalTotal'] = suma_capital2_deuda4

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor4 = [0, 0, 0]
    lst_deudor4[0] = dic_sd4g0
    lst_deudor4[1] = dic_sd4g1
    lst_deudor4[2] = dic_sd4g2
    # print(lst_deudor4)

    '''------------------------ DEUDA 5 ------------------------ '''

    # creo diccionarios para organizar la situacion de deuda en este caso 5, de acuerdo al tipo de garantia que tenga
    dic_sd5g0 = {'tipoDeuda': 'sinGarantias', 'capitalTotal': 0}
    dic_sd5g1 = {'tipoDeuda': 'garantiasA', 'capitalTotal': 0}
    dic_sd5g2 = {'tipoDeuda': 'garantiasB', 'capitalTotal': 0}

    suma_capital0_deuda5 = 0
    suma_capital1_deuda5 = 0
    suma_capital2_deuda5 = 0

    # recorra la lista de los de situacion de deuda 5
    for dic in lst_deudor5:
        for keys in dic:
            if keys == 'tipoGarantia':
                if dic['tipoGarantia'] == 0:
                    # sumo el capital de todos aquellos que no tengan garantia (0)
                    suma_capital0_deuda5 += dic['capital']
                if dic['tipoGarantia'] == 1:
                    # sumo el capital de todos aquellos que tengan garantia A (1)
                    suma_capital1_deuda5 += dic['capital']
                if dic['tipoGarantia'] == 2:
                    # sumo el capital de todos aquellos que tengan garantia B (2)
                    suma_capital2_deuda5 += dic['capital']

    # luego organizo todos los diccionarios con la suma de capital obtenida arriba
    dic_sd5g0['capitalTotal'] = suma_capital0_deuda5
    dic_sd5g1['capitalTotal'] = suma_capital1_deuda5
    dic_sd5g2['capitalTotal'] = suma_capital2_deuda5

    # re-organizo la lista de deudor con los diccionarios de arriba
    lst_deudor5 = [0, 0, 0]
    lst_deudor5[0] = dic_sd5g0
    lst_deudor5[1] = dic_sd5g1
    lst_deudor5[2] = dic_sd5g2
    # print(lst_deudor5)

    # lista data frame
    lista_deudores = []

    for x in lst_deudor1:
        lista_deudores.append(x)

    for x in lst_deudor2:
        lista_deudores.append(x)

    for x in lst_deudor3:
        lista_deudores.append(x)

    for x in lst_deudor4:
        lista_deudores.append(x)

    for x in lst_deudor5:
        lista_deudores.append(x)

    print(lista_deudores, 'lista deudores cartera comercial')

    # Lista para generar el excel
    datosDeudores = pd.DataFrame(lista_deudores)

    # creacion del excel
    datosDeudores.to_excel('cartera_comercialConsumo.xlsx')

    # ruta de nuestro archivo
    filesheet = "./cartera_comercialConsumo.xlsx"

    # creamos el objeto load_workbook
    wb = load_workbook(filesheet)

    # Seleccionamos el archivo
    sheet = wb.active

    # Deudor 1 - COlumna excel

    sheet['A1'] = 'Situacion deudor'
    sheet['A2'] = '1'
    sheet['A3'] = '1'
    sheet['A4'] = '1'

    sheet['A5'] = '2'
    sheet['A6'] = '2'
    sheet['A7'] = '2'

    sheet['A8'] = '3'
    sheet['A9'] = '3'
    sheet['A10'] = '3'

    sheet['A11'] = '4'
    sheet['A12'] = '4'
    sheet['A13'] = '4'

    sheet['A14'] = '5'
    sheet['A15'] = '5'
    sheet['A16'] = '5'
    # Guardamos el archivo con los cambios
    wb.save(filesheet)
# INICIO
def iniciarDataBase(clientes):

    # creacion de documento
    def typeDocument():
        value = r(0, 1)
        if value == 0:
            type = '011'
        else:
            type = '099'
        return type

# creacion de nombres aleaotorios
    def randomName():
        value = r(0, 1)
        if value == 1:
            sex = 'male'
        else:
            sex = 'female'

        name = names.get_full_name(gender=sex)
        return name

    # seleccion de tipo de operacion realizada por cada cliente
    def typeOperation():
        value = r(1, 37)
        if len(str(value)) == 2:
            operation = '0000' + str(value)
        else:
            operation = '00000' + str(value)
        return operation

    # creacion de documento de acuerdo el tipo de documento que tengas
    def randomDocument(valueTypeDocument):
        unSexDocument = str(r(12000000, 44000000))
        if valueTypeDocument == '011':
            secondValue = r(0, 1)
            if secondValue == 0:
                thirdValue = r(0, 1)
                if thirdValue == 0:
                    secondType = '30'
                    document = valueTypeDocument+secondType+unSexDocument
                    return document
                else:
                    secondType = '20'
                    document = valueTypeDocument+secondType+unSexDocument
                    return document
            elif secondValue == 1:
                thirdValue = r(0, 1)
                if thirdValue == 0:
                    secondType = '33'
                    document = valueTypeDocument+secondType+unSexDocument
                    return document
                else:
                    secondType = '27'
                    document = valueTypeDocument+secondType+unSexDocument
                    return document
        else:
            document = valueTypeDocument+unSexDocument
            return document

    # seleccion del codigo de garantia aleatorio que tendra el prestamo
    def randomGuaranteeCode():
        guaranteeCode = r(0, 33)
        return guaranteeCode

    # asigo el tipo de garantia de acuerdo al codigo de garantia ya establecido
    def typeGuarantee(guaranteeCode):
        if guaranteeCode == 0:
            typeValueGuarantee = 0
        elif guaranteeCode == 1 or guaranteeCode == 2 or guaranteeCode == 3 or guaranteeCode == 4 or guaranteeCode == 5 or guaranteeCode == 6 or guaranteeCode == 7 or guaranteeCode == 8 or guaranteeCode == 9 or guaranteeCode == 10 or guaranteeCode == 11 or guaranteeCode == 25 or guaranteeCode == 31:
            typeValueGuarantee = 2
        else:
            typeValueGuarantee = 1
        return typeValueGuarantee
    # seleccion del codigo de moneda

    def randomCurrencyCode():
        currencyCode = r(0, 1)
        return currencyCode

    # seleccion aleatoria de tipo de Cartera
    def randomTypeWallet():
        typeWallet = r(1, 3)
        return typeWallet

    # creacion de capital de operacion de acuerdo al tipo de cartera del cliente y la moneda
    def randomOperationCapital(typeWallet, currencyCode):
        if typeWallet == 1:
            operationCapital = r(50000, 30000000)
            if currencyCode == 0:
                return operationCapital
            else:
                return operationCapital//300
        elif typeWallet == 2:
            operationCapital = r(100000000, 500000000)
            if currencyCode == 0:
                return operationCapital
            else:
                return operationCapital//300
        else:
            operationCapital = r(30000000, 100000000)
            if currencyCode == 0:
                return operationCapital
            else:
                return operationCapital//300

    # asignacion de numero de operacion
    def operationNumber(operation):
        operationModify = operation[((len(operation))-2):(len(operation))]
        if operationModify == '01':
            opNumber = '131709'
        elif operationModify == '02':
            opNumber = '13712'
        elif operationModify == '03':
            opNumber = '13712'
        elif operationModify == '04':
            opNumber = '131718'
        elif operationModify == '05' or operationModify == '21':
            opNumber = '131708'
        elif operationModify == '06':
            opNumber = '131711'
        elif operationModify == '07':
            opNumber = '131713'
        elif operationModify == '08':
            opNumber = '131714'
        elif operationModify == '09':
            opNumber = '131731'
        elif operationModify == '10' or operationModify == '11' or operationModify == '12' or operationModify == '13' or operationModify == '14' or operationModify == '15' or operationModify == '16' or operationModify == '17':
            opNumber = '131742'
        elif operationModify == '18':
            opNumber = '131741'
        elif operationModify == '19':
            opNumber = '131738'
        elif operationModify == '20':
            opNumber = '131736'
        elif operationModify == '22':
            opNumber = '132735'
        elif operationModify == '23':
            opNumber = '721731'
        elif operationModify == '24':
            opNumber = '141701'
        elif operationModify == '25':
            opNumber = '150720'
        elif operationModify == '26':
            opNumber = '171131'
        elif operationModify == '27':
            opNumber = '131101'
        elif operationModify == '28':
            opNumber = '721735'
        elif operationModify == '30':
            opNumber = '131728'
        elif operationModify == '31':
            opNumber = '135799'
        elif operationModify == '32':
            opNumber = '161003'
        elif operationModify == '33':
            opNumber = '131744'
        elif operationModify == '34':
            opNumber = '131752'
        elif operationModify == '35':
            opNumber = '131748'
        elif operationModify == '36':
            opNumber = '131792'
        else:
            opNumber = '131792'
        return opNumber
    # asignacion de situacion de deudor de acuerdo al atraso

    def debtorAssignment():
        debt = r(0, 360)
        if debt >= 0 and debt <= 30:
            debtSituation = 1
        elif debt >= 31 and debt <= 90:
            debtSituation = 2
        elif debt >= 90 and debt <= 180:
            debtSituation = 3
        elif debt >= 181 and debt <= 360:
            debtSituation = 4
        else:
            debtSituation = 5
        return debtSituation

    # asignacion de interes a cobrar de acuerdo al capital
    def interestCharge(typeWallet, operationCapital):
        if typeWallet == 1:
            interest = round(operationCapital*0.1)
        elif typeWallet == 2:
            interest = round(operationCapital*0.2)
        elif typeWallet == 3:
            interest = operationCapital*0.25
        elif typeWallet == 4:
            interest = round(operationCapital*0.3)
        else:
            interest = round(operationCapital*0.35)
        return interest

    # creacion del dataframe
    keys = ['tipoDocumento', 'documento', 'nombreCompleto', 'operacion', 'codigoGarantia', 'tipoGarantia',
            'codigoMoneda', 'capitalOperacion', 'interesCobrar', 'clasificacionDeudor', 'numeroOperacion', 'tipoCartera']
    dataframe = []
    clients = clientes
    for x in range(clients):
        # creacion de diccionario con valores independientes
        dictionary = {'tipoDocumento': typeDocument(), 'documento': 0, 'nombreCompleto': randomName(), 'operacion': typeOperation(), 'codigoGarantia': randomGuaranteeCode(
        ), 'capitalOperacion': 0, 'interesCobrar': 0, 'tipoGarantia': 0, 'codigoMoneda': randomCurrencyCode(), 'clasificacionDeudor': debtorAssignment(), 'numeroOperacion': 0, 'tipoCartera': randomTypeWallet()}
        # asignando valores dependientes de los diccionarios
        dictionary['documento'] = randomDocument(dictionary['tipoDocumento'])
        dictionary['tipoGarantia'] = typeGuarantee(
            dictionary['codigoGarantia'])
        dictionary['capitalOperacion'] = randomOperationCapital(
            dictionary['tipoCartera'], dictionary['codigoMoneda'])
        dictionary['interesCobrar'] = interestCharge(
            dictionary['tipoCartera'], dictionary['capitalOperacion'])
        dictionary['numeroOperacion'] = operationNumber(
            dictionary['operacion'])
        # los coloco en el dataframe
        dataframe.append(dictionary)
        dictionary = {}

    # creacion de la base de datos
    print(dataframe)
    database = pd.DataFrame(dataframe)
    print(database)

    # creacion del archivo excel
    database.to_excel('database.xlsx')

    salida_uno(clientes)
    salida_dos(clientes)
    salida_tres(clientes)
    salida_errores()


def inicio():

    saludo = int(
        input(('Bienvenido al soft de gestion bancaria, coloque 1 para iniciar el programa: ')))
    while saludo != 1:
        print('Numero incorrecto')
        saludo = int(
            input(('Bienvenido al soft de gestion bancaria, coloque 1 para iniciar el programa: ')))

    if saludo == 1:
        clientes = int(
            input('Coloque la cantidad de clientes que desea que la base tenga: '))
        iniciarDataBase(clientes)


inicio()
